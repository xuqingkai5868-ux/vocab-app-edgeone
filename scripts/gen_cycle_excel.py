#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成 PET 词汇 6天周期分组 Excel
把每 6 天的单词和短语合并到一个 Sheet
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ===== 配置 =====
SCHEDULE_PATH = 'C:/CODE/vocab-app/pet_schedule_v2.json'
OUTPUT_PATH = 'C:/Users/xuqin/Desktop/PET词汇-6天周期分组.xlsx'
NUM_DAYS_PER_CYCLE = 6

# ===== 颜色 =====
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_FONT = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
WORD_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
PHRASE_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7'),
)
ALT_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

# ===== 读取数据 =====
with open(SCHEDULE_PATH, 'r', encoding='utf-8') as f:
    data = json.load(f)
schedule = data['schedule']

total_cycles = (len(schedule) + NUM_DAYS_PER_CYCLE - 1) // NUM_DAYS_PER_CYCLE

# ===== 创建 Workbook =====
wb = Workbook()
wb.remove(wb.active)

summary_rows = []

for cycle_idx in range(1, total_cycles + 1):
    start_day = (cycle_idx - 1) * NUM_DAYS_PER_CYCLE + 1
    end_day = min(cycle_idx * NUM_DAYS_PER_CYCLE, len(schedule))

    entries = []
    word_count = 0
    phrase_count = 0

    for d in range(start_day, end_day + 1):
        day_data = schedule[d - 1]
        for w in day_data.get('words', []):
            entries.append({
                'type': '单词',
                'word': w['word'],
                'pos': w.get('pos', ''),
                'meaning': w.get('meaning', ''),
                'day': d,
            })
            word_count += 1
        for p in day_data.get('phrases', []):
            entries.append({
                'type': '短语',
                'word': p['phrase'],
                'pos': '',
                'meaning': p.get('meaning', ''),
                'day': d,
                'source': p.get('source', ''),
                'assoc_w': p.get('associated_word', ''),
            })
            phrase_count += 1

    # Sheet 名
    sheet_name = f'周期{cycle_idx} Day{start_day}-{end_day}'
    if len(sheet_name) > 31:
        sheet_name = f'周期{cycle_idx}'
    ws = wb.create_sheet(title=sheet_name)

    # 列宽
    col_widths = {'A': 6, 'B': 26, 'C': 10, 'D': 34, 'E': 10, 'F': 8, 'G': 14, 'H': 14}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 标题行
    ws.merge_cells('A1:H1')
    c = ws.cell(row=1, column=1, value=f'PET 词汇 — 周期 {cycle_idx}（Day {start_day} - Day {end_day}）')
    c.font = Font(name='微软雅黑', bold=True, size=14, color='2F5496')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # 统计行
    ws.merge_cells('A2:H2')
    c = ws.cell(row=2, column=1, value=f'共 {len(entries)} 项（{word_count} 个单词 + {phrase_count} 个短语）')
    c.font = Font(name='微软雅黑', size=10, color='666666')
    c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 22

    # 表头
    headers = ['序号', '单词 / 短语', '词性', '释义', '所属Day', '类型', '来源', '关联词']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    # 数据行
    for idx, entry in enumerate(entries):
        row = 5 + idx
        is_word = entry['type'] == '单词'
        fill = WORD_FILL if is_word else PHRASE_FILL

        vals = [
            idx + 1,
            entry['word'],
            entry['pos'],
            entry['meaning'],
            f"Day {entry['day']}",
            entry['type'],
            entry.get('source', ''),
            entry.get('assoc_w', ''),
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = Font(name='微软雅黑', size=10)
            cell.fill = fill
            cell.border = THIN_BORDER
            center_cols = (1, 3, 5, 6)
            cell.alignment = Alignment(
                horizontal='center' if col_idx in center_cols else 'left',
                vertical='center'
            )

    summary_rows.append({
        'cycle': cycle_idx,
        'days': f'Day {start_day}-{end_day}',
        'words': word_count,
        'phrases': phrase_count,
        'total': len(entries),
    })

# ===== 汇总 Sheet =====
ws = wb.create_sheet(title='汇总', index=0)
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12

ws.merge_cells('A1:E1')
c = ws.cell(row=1, column=1, value='PET 词汇 — 6 天周期汇总')
c.font = Font(name='微软雅黑', bold=True, size=16, color='2F5496')
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

tw = sum(r['words'] for r in summary_rows)
tp = sum(r['phrases'] for r in summary_rows)
ws.merge_cells('A2:E2')
c = ws.cell(row=2, column=1, value=f'总计：{len(summary_rows)} 个周期 / {tw} 个单词 / {tp} 个短语')
c.font = Font(name='微软雅黑', size=10, color='666666')
c.alignment = Alignment(horizontal='center')

sh = ['周期', '天数范围', '单词数', '短语数', '总条目']
for col_idx, h in enumerate(sh, 1):
    cell = ws.cell(row=4, column=col_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER

for idx, sr in enumerate(summary_rows):
    row = 5 + idx
    vals = [sr['cycle'], sr['days'], sr['words'], sr['phrases'], sr['total']]
    for col_idx, val in enumerate(vals, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = Font(name='微软雅黑', size=10)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if idx % 2 == 1:
            cell.fill = ALT_FILL

# ===== 保存 =====
wb.save(OUTPUT_PATH)
print(f'Done: {OUTPUT_PATH}')
print(f'{total_cycles} cycles, {tw} words, {tp} phrases')
